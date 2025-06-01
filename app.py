import gradio as gr
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import math

# Define 3 sets : unit THB/Watt
THBperWatt_set1 = [None, 3.97, 5.60, 3.06, 3.10, 2.50, 0.60, 3.00, 7.00]#<100kWp
THBperWatt_set2 = [None, 3.50, 5.00, 2.90, 2.95, 2.30, 0.55, 2.80, 6.50]#guess
THBperWatt_set3 = [None, 3.00, 4.50, 2.70, 2.80, 2.10, 0.50, 2.50, 6.00]#guess:>1000kWp

def roundup(x):
    return math.ceil(x)

def select_THBperWatt(SolarInstallCapacity):
    if SolarInstallCapacity < 100:
        return THBperWatt_set1
    elif SolarInstallCapacity < 1000:
        return THBperWatt_set2
    else:
        return THBperWatt_set3

def analyse_fn(SolarInstallCapacity, customer_name_txt, customer_addr_txt, expire_date_txt, num5_txt, optimizer_checkbox, num6_txt, PannelPower, quotation_number_txt):
    # Use the new function to select the set
    THBperWatt = select_THBperWatt(SolarInstallCapacity)
    THBperkWp = [x * 1000 if x is not None else None for x in THBperWatt]
    pannel_capacity = 600/1000  # kWp per panel

    src = "form.xlsx"
    dst = "form_filled.xlsx"
    shutil.copy(src, dst)

    wb = load_workbook(dst)
    ws = wb.active

    # Fill header info
    ws["C8"] = "Customer: " + customer_name_txt
    ws["C8"].font = Font(name="TH SarabunPSK", size=12)
    ws["C9"] = "Bill To: " + customer_addr_txt
    ws["C9"].font = Font(name="TH SarabunPSK", size=12)
    ws["C28"] = "Until " + expire_date_txt
    ws["C28"].font = Font(name="TH SarabunPSK", size=12)
    ws["C33"] = num5_txt
    ws["C33"].font = Font(name="TH SarabunPSK", size=12)
    ws["C34"] = num6_txt
    ws["C34"].font = Font(name="TH SarabunPSK", size=12)
    ws["L5"] = quotation_number_txt
    ws["L5"].font = Font(name="TH SarabunPSK", size=12)

    # Update pannel_capacity from user input
    pannel_capacity = PannelPower / 1000  # kWp per panel

    n_pannel = roundup(SolarInstallCapacity / pannel_capacity)  # Number of panels needed
    # Fill item rows
    columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    items = [
        ["PV panels", roundup(SolarInstallCapacity * THBperkWp[1]/n_pannel), n_pannel, "Pannels"],
        ["Inverter and Accessories", SolarInstallCapacity * THBperkWp[2], 1, "LOT"]
    ]
    if optimizer_checkbox:
        items.append(["Optimizer/rapid shutdown", SolarInstallCapacity * THBperkWp[3], 1, "LOT"])

    items.append(["PV Support Structure", SolarInstallCapacity * THBperkWp[4], 1, "LOT"])
    items.append(["LABOUR INSTALATION", SolarInstallCapacity * THBperkWp[5], 1, "LOT"])
    items.append(["PERMIT AND LICENSE", SolarInstallCapacity * THBperkWp[6], 1, "LOT"])
    items.append(["Engineering and Construction", SolarInstallCapacity * THBperkWp[7], 1, "LOT"])
    items.append(["Other", SolarInstallCapacity * THBperkWp[8], 1, "LOT"])
    
    for idx, item in enumerate(items, start=1):
        row = 11 + idx  # C12, C13, ...
        values = [
            idx, item[0], None, None, item[1], "THB", item[2], item[3], item[1]*item[2], "THB"
        ]
        for col, val in zip(columns, values):
            ws[f"{col}{row}"] = val
            ws[f"{col}{row}"].font = Font(name="TH SarabunPSK", size=12)

    wb.save(dst)
    wb.close()
    return dst

def get_approver_list_xlsx(xlsx_path="approver_list.xlsx"):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    approvers = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0]:
            approvers.append(str(row[0]))  # เอาเฉพาะชื่อ
    wb.close()
    return approvers

def get_approver_positions_xlsx(xlsx_path="approver_list.xlsx"):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    positions = []
    seen = set()
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[1]:
            pos = str(row[1])
            if pos not in seen:
                positions.append(pos)
                seen.add(pos)
    wb.close()
    return positions

if __name__ == "__main__":
    approver_options = get_approver_list_xlsx()  # รายชื่อ
    position_options = get_approver_positions_xlsx()  # เฉพาะตำแหน่ง

    with gr.Blocks() as solar_price_estimator:
        gr.Markdown("ประมาณราคาโครงการโซล่าเซลล์")
        with gr.Tab("ประมาณราคา"):
            with gr.Row():
                SolarInstallCapacity = gr.Number(
                    label="กำลังผลิตติดตั้ง (kWp)",
                    step=1,
                    value=10
                )
                PannelPower = gr.Number(
                    label="กำลังแผง (Watt)",
                    step=10,
                    value=600
                )
            with gr.Row():
                customer_name_txt = gr.Textbox(label="ชื่อลูกค้า", value="- ชื่อลูกค้า -")
                customer_addr_txt = gr.Textbox(label="ที่อยู่ลูกค้า", value="- ที่อยู่ลูกค้า -")
            with gr.Row():
                optimizer_checkbox = gr.Checkbox(label="Optimizer/rapid shutdown", value=True)
            with gr.Row():
                expire_date_txt = gr.Textbox(label="วันหมดอายุใบเสนอราคา", value="- วันหมดอายุใบเสนอราคา -")
                approval_name_dd = gr.Dropdown(label="ชื่อผู้อนุมัติ", choices=approver_options, value=approver_options[0] if approver_options else "-")
                approval_position_dd = gr.Dropdown(label="ตำแหน่ง", choices=position_options, value=position_options[0] if position_options else "-")
            with gr.Row():
                quotation_number_txt = gr.Textbox(label="เลขที่ใบเสนอราคา", value="- Quotation No. -")
            with gr.Row():
                generate_report_btn = gr.Button("สร้างรายงาน")
            with gr.Row():
                download_file = gr.File(label="ดาวน์โหลดไฟล์ที่แก้ไขแล้ว")

            generate_report_btn.click(
                fn=analyse_fn,
                inputs=[
                    SolarInstallCapacity,
                    customer_name_txt,
                    customer_addr_txt,
                    expire_date_txt,
                    approval_name_dd,
                    optimizer_checkbox,
                    approval_position_dd,
                    PannelPower,
                    quotation_number_txt  # เพิ่มตรงนี้
                ],
                outputs=download_file
            )

    solar_price_estimator.launch()